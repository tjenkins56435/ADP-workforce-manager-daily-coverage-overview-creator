#!/usr/bin/env python3
"""
Daily Playbook Creator (DCO Creator)
Imports an ADP "Employee Schedule - Weekly" Excel report,
lets the user pick a day, assign zone colors, and generates
a formatted Daily Playbook Excel file.
"""

import json
import os
import re
import tkinter as tk
from tkinter import ttk, colorchooser, filedialog, messagebox
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    load_workbook = None
    Workbook = None

# ---------------------------------------------------------------------------
# Config persistence
# ---------------------------------------------------------------------------
CONFIG_PATH = Path(__file__).parent / "config.json"
DEFAULT_ZONES = [
    {"name": "Adults", "color": "#FF0000"},
    {"name": "Kids/Footwear", "color": "#FFD700"},
    {"name": "Cashiers", "color": "#00C853"},
    {"name": "Replenishment/Refill", "color": "#9C27B0"},
    {"name": "Shipment", "color": "#2196F3"},
    {"name": "Operation", "color": "#FF922B"},
    {"name": "Fitting Rooms", "color": "#FF7F50"},
]

DAY_NAMES = ["Sunday", "Monday", "Tuesday", "Wednesday",
             "Thursday", "Friday", "Saturday"]
DAY_ABBREVS = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]


def load_config():
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {
        "zones": DEFAULT_ZONES,
        "last_output_dir": str(Path.home() / "Desktop"),
        "last_import_dir": str(Path.home() / "Downloads"),
    }


def save_config(cfg):
    try:
        with open(CONFIG_PATH, "w") as f:
            json.dump(cfg, f, indent=2)
    except IOError as e:
        print(f"Warning: could not save config: {e}")


# ---------------------------------------------------------------------------
# Time helpers
# ---------------------------------------------------------------------------

def _text_color_for_bg(hex_color):
    """Return '#000000' or '#FFFFFF' for readable text on the given background."""
    hex_color = hex_color.lstrip("#")
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    return "#000000" if luminance > 0.5 else "#FFFFFF"


def parse_time_range(text):
    """Parse '10:00 AM - 2:00 PM' into ((hour, min), (hour, min)) tuple."""
    if not text or not isinstance(text, str):
        return None
    text = text.strip()
    m = re.match(
        r'(\d{1,2}(?::\d{2})?)\s*(AM|PM)\s*[-–]\s*(\d{1,2}(?::\d{2})?)\s*(AM|PM)',
        text, re.IGNORECASE
    )
    if not m:
        return None
    start = _parse_hm(m.group(1), m.group(2))
    end = _parse_hm(m.group(3), m.group(4))
    if start and end:
        return (start, end)
    return None


def _parse_hm(time_str, ampm):
    """Parse '10:00' or '10' with AM/PM into (hour, minute)."""
    ampm = ampm.upper()
    if ":" in time_str:
        parts = time_str.split(":")
        hour, minute = int(parts[0]), int(parts[1])
    else:
        hour, minute = int(time_str), 0
    if ampm == "PM" and hour != 12:
        hour += 12
    if ampm == "AM" and hour == 12:
        hour = 0
    return (hour, minute)


def format_time_short(hour, minute):
    """Format (hour, minute) as compact string like '9A', '930A', '12P'."""
    if hour == 0:
        display_hour, suffix = 12, "A"
    elif hour < 12:
        display_hour, suffix = hour, "A"
    elif hour == 12:
        display_hour, suffix = 12, "P"
    else:
        display_hour, suffix = hour - 12, "P"
    if minute == 0:
        return f"{display_hour}{suffix}"
    return f"{display_hour}{minute:02d}{suffix}"


def generate_time_slots(start_hour, end_hour):
    """Generate list of (hour, minute) tuples for half-hour slots."""
    slots = []
    h, m = start_hour, 0
    while h < end_hour or (h == end_hour and m == 0):
        slots.append((h, m))
        m += 30
        if m >= 60:
            m = 0
            h += 1
    return slots


def _get_zone_for_slot(emp, slot_h, slot_m):
    """Return the zone name for a given time slot, or '' if not covered.

    Multi-zone: checks zone_segments and shift_segments to skip break gaps.
    Single zone: uses overall shift_start/shift_end range (colors continuously).
    """
    slot_min = slot_h * 60 + slot_m

    # Multi-zone: use zone_segments; fill break gaps with preceding zone
    zone_segs = emp.get("zone_segments", [])
    if zone_segs:
        # Check if slot falls within a zone segment directly
        for seg in zone_segs:
            seg_start = seg["start"][0] * 60 + seg["start"][1]
            seg_end = seg["end"][0] * 60 + seg["end"][1]
            if seg_start <= slot_min < seg_end:
                return seg["zone"]
        # Slot is in a break gap — extend the preceding zone segment's color
        shift_start = emp.get("shift_start")
        shift_end = emp.get("shift_end")
        if shift_start and shift_end:
            start_min = shift_start[0] * 60 + shift_start[1]
            end_min = shift_end[0] * 60 + shift_end[1]
            if start_min <= slot_min < end_min:
                # Find the last zone segment that ended at or before this slot
                best = None
                for seg in zone_segs:
                    seg_end = seg["end"][0] * 60 + seg["end"][1]
                    if seg_end <= slot_min:
                        best = seg["zone"]
                if best:
                    return best
        return ""

    # Single zone: color the full shift range continuously
    shift_start = emp.get("shift_start")
    shift_end = emp.get("shift_end")
    if shift_start and shift_end:
        start_min = shift_start[0] * 60 + shift_start[1]
        end_min = shift_end[0] * 60 + shift_end[1]
        if start_min <= slot_min < end_min:
            return emp.get("zone", "")
    return ""


# ---------------------------------------------------------------------------
# ADP Excel Report Parser
# ---------------------------------------------------------------------------

def parse_adp_report(filepath):
    """Parse an ADP Employee Schedule Weekly report.

    Returns:
        {
            "time_period": str,
            "days": {col_index: {"abbrev": "Sun", "date": datetime}, ...},
            "employees": [
                {
                    "name": str,
                    "job": str,
                    "department": str,
                    "shifts_by_day": {
                        col_index: [((start_h, start_m), (end_h, end_m)), ...],
                        ...
                    }
                }, ...
            ]
        }
    """
    if load_workbook is None:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    result = {
        "time_period": "",
        "days": {},
        "employees": [],
    }

    # Read all cell values into a grid for easier processing
    max_row = ws.max_row
    max_col = ws.max_column
    grid = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col,
                            values_only=True):
        grid.append(list(row))

    if not grid:
        return result

    # Find time period (row 2, col E typically)
    for r in range(min(5, len(grid))):
        for c in range(len(grid[r])):
            val = grid[r][c]
            if isinstance(val, str) and re.match(r'\d+/\d+/\d+\s*-\s*\d+/\d+/\d+', val):
                result["time_period"] = val
                break

    # Parse the file section by section
    current_dept = ""
    day_columns = {}  # col_index -> {"abbrev": "Sun", "date": datetime}
    current_employee = None
    r = 0

    while r < len(grid):
        row = grid[r]
        cell_a = row[0] if row[0] else ""

        # Department header: long path string
        if isinstance(cell_a, str) and cell_a.startswith("Under Armour/"):
            parts = cell_a.split("/")
            current_dept = parts[-1] if parts else cell_a
            r += 1
            continue

        # Header row: "Employee" in col A, day abbreviations
        if cell_a == "Employee":
            # Next row should have dates - find day columns
            day_columns = {}
            if r + 1 < len(grid):
                date_row = grid[r + 1]
                header_row = row
                for c in range(len(header_row)):
                    hval = header_row[c]
                    if isinstance(hval, str) and hval in DAY_ABBREVS:
                        abbrev = hval
                        # Check the date row for a datetime
                        date_val = date_row[c] if c < len(date_row) else None
                        day_columns[c] = {
                            "abbrev": abbrev,
                            "date": date_val if isinstance(date_val, datetime) else None,
                        }
                r += 2  # skip header + date row
                continue
            r += 1
            continue

        # Employee name row: name in col A, job in col G (index 6)
        if isinstance(cell_a, str) and cell_a and "," in cell_a and not cell_a.startswith("Under"):
            # Looks like "Last, First" format
            job = row[6] if len(row) > 6 and row[6] else ""

            # Collect shifts for this employee from this row
            shifts_by_day = {}
            for col_idx, day_info in day_columns.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    tr = parse_time_range(str(val) if val else "")
                    if tr:
                        shifts_by_day.setdefault(col_idx, []).append(tr)

            current_employee = {
                "name": cell_a.strip(),
                "job": str(job).strip() if job else "",
                "department": current_dept,
                "shifts_by_day": shifts_by_day,
            }
            result["employees"].append(current_employee)
            r += 1
            continue

        # Continuation row (no name in col A): additional shifts for current employee
        if current_employee and not cell_a:
            has_shift = False
            for col_idx, day_info in day_columns.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    tr = parse_time_range(str(val) if val else "")
                    if tr:
                        current_employee["shifts_by_day"].setdefault(col_idx, []).append(tr)
                        has_shift = True
            if not has_shift:
                # Could be a job-type label row or blank - skip
                pass
            r += 1
            continue

        r += 1

    # Build the global days mapping from all discovered day_columns
    if not result["days"] and day_columns:
        result["days"] = day_columns

    return result


def extract_day_schedule(parsed, target_col):
    """Extract employees working on a specific day column.

    Returns list of employee dicts ready for the GUI:
        [{name, shift_text, shift_segments, break_text, job, zone}, ...]
    """
    employees = []
    for emp in parsed["employees"]:
        segments = emp["shifts_by_day"].get(target_col, [])
        if not segments:
            continue

        # Sort segments by start time
        segments.sort(key=lambda s: s[0][0] * 60 + s[0][1])

        # Build shift text
        shift_parts = []
        for (sh, sm), (eh, em) in segments:
            shift_parts.append(f"{format_time_short(sh, sm)}-{format_time_short(eh, em)}")
        shift_text = " / ".join(shift_parts)

        # Infer break from gap between segments
        break_text = ""
        if len(segments) >= 2:
            # Break is between end of first segment and start of second
            _, end1 = segments[0]
            start2, _ = segments[1]
            break_text = format_time_short(*end1)

        # Overall start/end for color bar
        overall_start = segments[0][0]
        overall_end = segments[-1][1]

        employees.append({
            "name": emp["name"],
            "job": emp.get("job", ""),
            "shift_text": shift_text,
            "shift_segments": segments,
            "shift_start": overall_start,
            "shift_end": overall_end,
            "break_text": break_text,
            "zone": "",
            "zone_segments": [],
        })

    return employees


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def generate_excel(employees, zones, day_name, date_str, output_path,
                   start_hour=9, end_hour=21):
    """Generate the Daily Playbook Excel file."""
    if Workbook is None:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Playbook"

    # Build zone color lookup
    zone_colors = {}
    for z in zones:
        zone_colors[z["name"]] = z["color"].lstrip("#")

    # Determine time range from employee shifts
    if employees:
        all_starts = [e["shift_start"] for e in employees if e.get("shift_start")]
        all_ends = [e["shift_end"] for e in employees if e.get("shift_end")]
        if all_starts:
            earliest = min(h for h, m in all_starts)
            start_hour = min(start_hour, earliest)
        if all_ends:
            latest = max(h + (1 if m > 0 else 0) for h, m in all_ends)
            end_hour = max(end_hour, latest)

    time_slots = generate_time_slots(start_hour, end_hour)
    time_col_start = 4  # Column D onwards

    # Styles
    header_font = Font(name="Calibri", bold=True, size=14)
    col_header_font = Font(name="Calibri", bold=True, size=10)
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Focuses column (2 merged columns after time slots)
    focuses_col_start = time_col_start + len(time_slots)

    # Row 1: Day and date header
    total_cols = focuses_col_start + 1  # includes both Focuses columns
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=total_cols)
    header_cell = ws.cell(row=1, column=1, value=f"{day_name} {date_str}")
    header_cell.font = header_font
    header_cell.alignment = center_align

    # Row 3: Column headers
    headers = ["TEAMMATE NAME", "SHIFT", "BREAK"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=3, column=i + 1, value=h)
        cell.font = col_header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                fill_type="solid")

    # Time slot headers
    for i, (h, m) in enumerate(time_slots):
        label = format_time_short(h, m)
        cell = ws.cell(row=3, column=time_col_start + i, value=label)
        cell.font = Font(name="Calibri", bold=True, size=8)
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                fill_type="solid")

    # Focuses header (merged across 2 columns)
    ws.merge_cells(start_row=3, start_column=focuses_col_start,
                   end_row=3, end_column=focuses_col_start + 1)
    focuses_header = ws.cell(row=3, column=focuses_col_start, value="FOCUSES")
    focuses_header.font = col_header_font
    focuses_header.alignment = center_align
    focuses_header.border = thin_border
    focuses_header.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                      fill_type="solid")
    # Border on the second merged cell too
    ws.cell(row=3, column=focuses_col_start + 1).border = thin_border

    # Pre-build a fill cache for zone colors
    fill_cache = {}
    for zname, zhex in zone_colors.items():
        fill_cache[zname] = PatternFill(start_color=zhex, end_color=zhex,
                                        fill_type="solid")

    # Employee rows
    for row_idx, emp in enumerate(employees):
        row = row_idx + 4

        # Name
        name_cell = ws.cell(row=row, column=1, value=emp["name"])
        name_cell.font = data_font
        name_cell.alignment = left_align
        name_cell.border = thin_border

        # Shift
        ws.cell(row=row, column=2, value=emp.get("shift_text", "")).font = data_font
        ws.cell(row=row, column=2).alignment = center_align
        ws.cell(row=row, column=2).border = thin_border

        # Break
        ws.cell(row=row, column=3, value=emp.get("break_text", "")).font = data_font
        ws.cell(row=row, column=3).alignment = center_align
        ws.cell(row=row, column=3).border = thin_border

        # Focuses (zone name, merged across 2 columns)
        zone_name = emp.get("zone", "")
        ws.merge_cells(start_row=row, start_column=focuses_col_start,
                       end_row=row, end_column=focuses_col_start + 1)
        focuses_cell = ws.cell(row=row, column=focuses_col_start, value=zone_name)
        focuses_cell.font = data_font
        focuses_cell.alignment = center_align
        focuses_cell.border = thin_border
        ws.cell(row=row, column=focuses_col_start + 1).border = thin_border

        # Color the time slot cells using per-slot zone lookup
        for i, (slot_h, slot_m) in enumerate(time_slots):
            cell = ws.cell(row=row, column=time_col_start + i)
            cell.border = thin_border

            slot_zone = _get_zone_for_slot(emp, slot_h, slot_m)
            if slot_zone and slot_zone in fill_cache:
                cell.fill = fill_cache[slot_zone]

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    for i in range(len(time_slots)):
        col_letter = get_column_letter(time_col_start + i)
        ws.column_dimensions[col_letter].width = 5
    for i in range(2):
        col_letter = get_column_letter(focuses_col_start + i)
        ws.column_dimensions[col_letter].width = 12

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "3:3"

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# GUI Application
# ---------------------------------------------------------------------------

class DCOCreatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Daily Playbook Creator")
        self.root.geometry("950x720")
        self.root.minsize(800, 600)

        self.config = load_config()
        self.employees = []
        self.zones = list(self.config.get("zones", DEFAULT_ZONES))
        self.parsed_report = None  # Holds full parsed ADP data
        self.available_days = {}   # col_index -> {abbrev, date}

        self._build_gui()
        self._refresh_zone_list()

    # ----- GUI Construction -----

    def _build_gui(self):
        main = ttk.Frame(self.root, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        self._build_import_panel(main)
        self._build_day_panel(main)
        self._build_zone_panel(main)
        self._build_employee_panel(main)
        self._build_action_panel(main)

    def _build_import_panel(self, parent):
        frame = ttk.LabelFrame(parent, text="Import ADP Report", padding=8)
        frame.pack(fill=tk.X, pady=(0, 5))

        row = ttk.Frame(frame)
        row.pack(fill=tk.X)

        ttk.Button(row, text="Open ADP Report (.xlsx)",
                   command=self._import_report).pack(side=tk.LEFT)

        self.file_var = tk.StringVar(value="No file loaded")
        ttk.Label(row, textvariable=self.file_var,
                  foreground="gray").pack(side=tk.LEFT, padx=(10, 0))

        ttk.Label(frame,
                  text="Export the \"Employee Schedule - Weekly\" report from "
                       "ADP as .xlsx, then open it here.",
                  foreground="gray").pack(anchor=tk.W, pady=(5, 0))

    def _build_day_panel(self, parent):
        frame = ttk.LabelFrame(parent, text="Select Day", padding=8)
        frame.pack(fill=tk.X, pady=(0, 5))

        row = ttk.Frame(frame)
        row.pack(fill=tk.X)

        ttk.Label(row, text="Day:").pack(side=tk.LEFT)
        self.day_combo_var = tk.StringVar()
        self.day_combo = ttk.Combobox(row, textvariable=self.day_combo_var,
                                      width=30, state="readonly")
        self.day_combo.pack(side=tk.LEFT, padx=(5, 15))
        self.day_combo.bind("<<ComboboxSelected>>", self._on_day_selected)

        ttk.Button(row, text="Load Day",
                   command=self._load_day).pack(side=tk.LEFT)

        self.day_info_var = tk.StringVar()
        ttk.Label(row, textvariable=self.day_info_var,
                  foreground="blue").pack(side=tk.LEFT, padx=(15, 0))

    def _build_zone_panel(self, parent):
        frame = ttk.LabelFrame(parent, text="Zone Configuration", padding=8)
        frame.pack(fill=tk.X, pady=(0, 5))

        btn_row = ttk.Frame(frame)
        btn_row.pack(fill=tk.X, pady=(0, 5))
        ttk.Button(btn_row, text="+ Add Zone",
                   command=self._add_zone).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_row, text="Edit Zone",
                   command=self._edit_zone).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_row, text="Delete Zone",
                   command=self._delete_zone).pack(side=tk.LEFT)

        cols = ("name", "color")
        self.zone_tree = ttk.Treeview(frame, columns=cols, show="headings",
                                      height=4)
        self.zone_tree.heading("name", text="Zone Name")
        self.zone_tree.heading("color", text="Color")
        self.zone_tree.column("name", width=200)
        self.zone_tree.column("color", width=120)
        self.zone_tree.pack(fill=tk.X)

    def _build_employee_panel(self, parent):
        frame = ttk.LabelFrame(parent, text="Employee Schedule", padding=8)
        frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        cols = ("name", "job", "shift", "break_time", "zone")
        self.emp_tree = ttk.Treeview(frame, columns=cols, show="headings",
                                     selectmode="browse")
        self.emp_tree.heading("name", text="Name")
        self.emp_tree.heading("job", text="Job")
        self.emp_tree.heading("shift", text="Shift")
        self.emp_tree.heading("break_time", text="Break")
        self.emp_tree.heading("zone", text="Zone")
        self.emp_tree.column("name", width=160)
        self.emp_tree.column("job", width=90)
        self.emp_tree.column("shift", width=150)
        self.emp_tree.column("break_time", width=70)
        self.emp_tree.column("zone", width=150)

        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL,
                                  command=self.emp_tree.yview)
        self.emp_tree.configure(yscrollcommand=scrollbar.set)
        self.emp_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.LEFT, fill=tk.Y)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(5, 0))

        ttk.Button(btn_frame, text="\u25B2 Up", width=10,
                   command=self._move_up).pack(pady=2)
        ttk.Button(btn_frame, text="\u25BC Down", width=10,
                   command=self._move_down).pack(pady=2)
        ttk.Button(btn_frame, text="Set Zone", width=10,
                   command=self._set_zone).pack(pady=2)
        ttk.Button(btn_frame, text="Set All Zones", width=10,
                   command=self._set_all_zones).pack(pady=2)
        ttk.Button(btn_frame, text="Edit", width=10,
                   command=self._edit_employee).pack(pady=2)
        ttk.Button(btn_frame, text="Add Manual", width=10,
                   command=self._add_manual_entry).pack(pady=2)
        ttk.Button(btn_frame, text="Delete", width=10,
                   command=self._delete_employee).pack(pady=2)

        self.emp_tree.bind("<Double-1>", lambda e: self._set_zone())

    def _build_action_panel(self, parent):
        frame = ttk.Frame(parent, padding=(0, 5))
        frame.pack(fill=tk.X)

        ttk.Button(frame, text="Generate Excel",
                   command=self._generate_excel).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame, text="Preview",
                   command=self._preview).pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(frame, text="Output folder:").pack(side=tk.LEFT, padx=(10, 5))
        self.output_var = tk.StringVar(
            value=self.config.get("last_output_dir",
                                  str(Path.home() / "Desktop")))
        ttk.Entry(frame, textvariable=self.output_var, width=35).pack(
            side=tk.LEFT, padx=(0, 5))
        ttk.Button(frame, text="Browse",
                   command=self._browse_output).pack(side=tk.LEFT)

    # ----- Import -----

    def _import_report(self):
        init_dir = self.config.get("last_import_dir",
                                   str(Path.home() / "Downloads"))
        path = filedialog.askopenfilename(
            initialdir=init_dir,
            title="Select ADP Report",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return

        try:
            self.parsed_report = parse_adp_report(path)
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to parse report:\n{e}")
            return

        self.config["last_import_dir"] = str(Path(path).parent)
        save_config(self.config)

        self.file_var.set(Path(path).name)

        # Populate day selector
        days = self.parsed_report.get("days", {})
        if not days:
            messagebox.showwarning("No Days Found",
                                   "Could not find day columns in the report.")
            return

        self.available_days = days
        options = []
        for col_idx in sorted(days.keys()):
            d = days[col_idx]
            abbrev = d["abbrev"]
            date_obj = d.get("date")
            if date_obj and isinstance(date_obj, datetime):
                label = f"{abbrev} {date_obj.month}/{date_obj.day}"
            else:
                label = abbrev
            options.append((col_idx, label))

        self.day_options = options
        self.day_combo["values"] = [label for _, label in options]
        if options:
            self.day_combo.current(0)

        emp_count = len(self.parsed_report.get("employees", []))
        period = self.parsed_report.get("time_period", "")
        messagebox.showinfo("Import Successful",
                            f"Loaded {emp_count} employees.\n"
                            f"Period: {period}\n\n"
                            f"Select a day and click 'Load Day'.")

    def _on_day_selected(self, event=None):
        pass  # Day is loaded explicitly via Load Day button

    def _load_day(self):
        if not self.parsed_report:
            messagebox.showinfo("No Report", "Import an ADP report first.")
            return

        idx = self.day_combo.current()
        if idx < 0 or idx >= len(self.day_options):
            messagebox.showinfo("Select Day", "Pick a day from the dropdown.")
            return

        col_idx, label = self.day_options[idx]
        self.employees = extract_day_schedule(self.parsed_report, col_idx)
        self._refresh_employee_list()

        day_info = self.available_days.get(col_idx, {})
        date_obj = day_info.get("date")
        if date_obj and isinstance(date_obj, datetime):
            self._current_day_name = DAY_NAMES[date_obj.weekday()]
            # Python weekday: 0=Mon. We need to match properly.
            # datetime.weekday(): Mon=0..Sun=6
            # For display, use strftime
            self._current_day_name = date_obj.strftime("%A")
            self._current_date_str = f"{date_obj.month}/{date_obj.day}"
        else:
            self._current_day_name = day_info.get("abbrev", "")
            self._current_date_str = ""

        self.day_info_var.set(
            f"{len(self.employees)} employees scheduled for "
            f"{self._current_day_name} {self._current_date_str}")

        # Auto-open zone assignment after loading a day
        if self.employees:
            self.root.after(100, lambda: self._open_zone_assignment(
                start_index=0, mode="sequential"))

    # ----- Zone Management -----

    def _refresh_zone_list(self):
        for item in self.zone_tree.get_children():
            self.zone_tree.delete(item)
        for z in self.zones:
            self.zone_tree.insert("", tk.END, values=(z["name"], z["color"]))

    def _add_zone(self):
        dialog = ZoneDialog(self.root, "Add Zone")
        if dialog.result:
            self.zones.append(dialog.result)
            self._refresh_zone_list()
            self._save_zones()

    def _edit_zone(self):
        sel = self.zone_tree.selection()
        if not sel:
            messagebox.showinfo("Edit Zone", "Select a zone to edit.")
            return
        idx = self.zone_tree.index(sel[0])
        dialog = ZoneDialog(self.root, "Edit Zone", self.zones[idx])
        if dialog.result:
            self.zones[idx] = dialog.result
            self._refresh_zone_list()
            self._save_zones()

    def _delete_zone(self):
        sel = self.zone_tree.selection()
        if not sel:
            messagebox.showinfo("Delete Zone", "Select a zone to delete.")
            return
        idx = self.zone_tree.index(sel[0])
        name = self.zones[idx]["name"]
        if messagebox.askyesno("Delete Zone", f"Delete zone '{name}'?"):
            self.zones.pop(idx)
            self._refresh_zone_list()
            self._save_zones()

    def _save_zones(self):
        self.config["zones"] = self.zones
        save_config(self.config)

    # ----- Employee List -----

    def _refresh_employee_list(self):
        for item in self.emp_tree.get_children():
            self.emp_tree.delete(item)
        for emp in self.employees:
            self.emp_tree.insert("", tk.END, values=(
                emp["name"],
                emp.get("job", ""),
                emp.get("shift_text", ""),
                emp.get("break_text", ""),
                emp.get("zone", ""),
            ))

    def _get_selected_index(self):
        sel = self.emp_tree.selection()
        if not sel:
            return None
        return self.emp_tree.index(sel[0])

    def _move_up(self):
        idx = self._get_selected_index()
        if idx is None or idx == 0:
            return
        self.employees[idx], self.employees[idx - 1] = (
            self.employees[idx - 1], self.employees[idx])
        self._refresh_employee_list()
        self.emp_tree.selection_set(self.emp_tree.get_children()[idx - 1])

    def _move_down(self):
        idx = self._get_selected_index()
        if idx is None or idx >= len(self.employees) - 1:
            return
        self.employees[idx], self.employees[idx + 1] = (
            self.employees[idx + 1], self.employees[idx])
        self._refresh_employee_list()
        self.emp_tree.selection_set(self.emp_tree.get_children()[idx + 1])

    def _open_zone_assignment(self, start_index=0, mode="sequential"):
        """Open the zone assignment dialog."""
        if not self.employees:
            messagebox.showinfo("No Employees", "Load a day first.")
            return
        if not self.zones:
            messagebox.showinfo("No Zones", "Add zones first.")
            return
        ZoneAssignmentDialog(self.root, self.employees, self.zones,
                             start_index=start_index, mode=mode)
        self._refresh_employee_list()

    def _set_zone(self):
        idx = self._get_selected_index()
        if idx is None:
            messagebox.showinfo("Set Zone", "Select an employee first.")
            return
        self._open_zone_assignment(start_index=idx, mode="single")

    def _set_all_zones(self):
        """Open sequential zone assignment for all employees."""
        self._open_zone_assignment(start_index=0, mode="sequential")

    def _edit_employee(self):
        idx = self._get_selected_index()
        if idx is None:
            messagebox.showinfo("Edit", "Select an employee first.")
            return
        dialog = EmployeeEditDialog(self.root, self.employees[idx])
        if dialog.result:
            self.employees[idx] = dialog.result
            self._refresh_employee_list()

    def _add_manual_entry(self):
        dialog = EmployeeEditDialog(self.root, None)
        if dialog.result:
            self.employees.append(dialog.result)
            self._refresh_employee_list()

    def _delete_employee(self):
        idx = self._get_selected_index()
        if idx is None:
            messagebox.showinfo("Delete", "Select an employee first.")
            return
        name = self.employees[idx]["name"]
        if messagebox.askyesno("Delete", f"Remove '{name}'?"):
            self.employees.pop(idx)
            self._refresh_employee_list()

    # ----- Excel Generation -----

    def _generate_excel(self):
        if not self.employees:
            messagebox.showinfo("No Data", "No employees loaded.")
            return

        output_dir = self.output_var.get().strip()
        if not output_dir or not os.path.isdir(output_dir):
            messagebox.showerror("Invalid Path", "Select a valid output folder.")
            return

        day_name = getattr(self, "_current_day_name", "")
        date_str = getattr(self, "_current_date_str", "")
        if not day_name:
            day_name = datetime.now().strftime("%A")
            date_str = f"{datetime.now().month}/{datetime.now().day}"

        safe_date = date_str.replace("/", "-") if date_str else "output"
        filename = f"Daily_Playbook_{safe_date}.xlsx"
        output_path = os.path.join(output_dir, filename)

        try:
            generate_excel(self.employees, self.zones,
                           day_name, date_str, output_path)
            self.config["last_output_dir"] = output_dir
            save_config(self.config)
            messagebox.showinfo("Success", f"Saved to:\n{output_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate Excel:\n{e}")

    def _preview(self):
        if not self.employees:
            messagebox.showinfo("No Data", "No employees to preview.")
            return

        day_name = getattr(self, "_current_day_name", "")
        date_str = getattr(self, "_current_date_str", "")
        lines = [f"  {day_name} {date_str}", ""]
        lines.append(f"  {'NAME':<22} {'SHIFT':<22} {'BREAK':<8} {'ZONE':<25}")
        lines.append("  " + "-" * 79)
        for emp in self.employees:
            lines.append(
                f"  {emp['name']:<22} {emp.get('shift_text',''):<22} "
                f"{emp.get('break_text',''):<8} {emp.get('zone',''):<25}")
        PreviewDialog(self.root, "\n".join(lines))

    def _browse_output(self):
        d = filedialog.askdirectory(initialdir=self.output_var.get())
        if d:
            self.output_var.set(d)


# ---------------------------------------------------------------------------
# Dialogs
# ---------------------------------------------------------------------------

class ZoneDialog:
    def __init__(self, parent, title, zone=None):
        self.result = None
        self.color = zone["color"] if zone else "#FF6B6B"

        self.win = tk.Toplevel(parent)
        self.win.title(title)
        self.win.geometry("300x150")
        self.win.transient(parent)
        self.win.grab_set()

        ttk.Label(self.win, text="Zone Name:").pack(padx=10, pady=(10, 2),
                                                     anchor=tk.W)
        self.name_var = tk.StringVar(value=zone["name"] if zone else "")
        ttk.Entry(self.win, textvariable=self.name_var, width=30).pack(
            padx=10, anchor=tk.W)

        color_frame = ttk.Frame(self.win)
        color_frame.pack(padx=10, pady=5, anchor=tk.W)
        ttk.Label(color_frame, text="Color:").pack(side=tk.LEFT)
        self.color_btn = tk.Button(color_frame, text="  ", bg=self.color,
                                   width=4, command=self._pick_color)
        self.color_btn.pack(side=tk.LEFT, padx=5)
        self.color_label = ttk.Label(color_frame, text=self.color)
        self.color_label.pack(side=tk.LEFT)

        btn_frame = ttk.Frame(self.win)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="OK", command=self._ok).pack(
            side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=self.win.destroy).pack(
            side=tk.LEFT, padx=5)

        self.win.wait_window()

    def _pick_color(self):
        _, hex_color = colorchooser.askcolor(initialcolor=self.color,
                                              title="Pick Zone Color")
        if hex_color:
            self.color = hex_color
            self.color_btn.configure(bg=hex_color)
            self.color_label.configure(text=hex_color)

    def _ok(self):
        name = self.name_var.get().strip()
        if not name:
            messagebox.showwarning("Warning", "Zone name cannot be empty.",
                                   parent=self.win)
            return
        self.result = {"name": name, "color": self.color}
        self.win.destroy()


class ZoneAssignmentDialog:
    """Large zone-assignment dialog with colored buttons.

    Modes:
        "sequential" - auto-advance through all employees
        "single"     - pick one zone then close
    """

    def __init__(self, parent, employees, zones, start_index=0,
                 mode="sequential"):
        self.employees = employees
        self.zones = zones
        self.mode = mode
        self.current_index = start_index
        self.cancelled = False

        self.win = tk.Toplevel(parent)
        self.win.title("Zone Assignment")
        self.win.geometry("600x500")
        self.win.transient(parent)
        self.win.grab_set()

        self._build_ui()
        self._show_employee()

        self.win.bind("<Escape>", lambda e: self._cancel())
        self.win.bind("<Left>", lambda e: self._prev())
        self.win.bind("<Right>", lambda e: self._next())
        for i in range(1, 10):
            self.win.bind(str(i), lambda e, idx=i: self._zone_by_number(idx))

        self.win.protocol("WM_DELETE_WINDOW", self._cancel)
        self.win.wait_window()

    def _build_ui(self):
        # Header
        header_frame = ttk.Frame(self.win, padding=(15, 10, 15, 5))
        header_frame.pack(fill=tk.X)

        self.name_label = ttk.Label(header_frame, text="",
                                    font=("Helvetica", 24, "bold"))
        self.name_label.pack(anchor=tk.W)

        self.info_label = ttk.Label(header_frame, text="",
                                    font=("Helvetica", 16),
                                    foreground="gray")
        self.info_label.pack(anchor=tk.W)

        # Progress
        progress_frame = ttk.Frame(self.win, padding=(15, 5))
        progress_frame.pack(fill=tk.X)

        self.progress_label = ttk.Label(progress_frame, text="")
        self.progress_label.pack(anchor=tk.W)

        self.progress_bar = ttk.Progressbar(progress_frame,
                                            mode="determinate", length=300)
        self.progress_bar.pack(fill=tk.X, pady=(2, 0))

        # Navigation
        nav_frame = ttk.Frame(self.win, padding=(15, 5))
        nav_frame.pack(fill=tk.X)

        self.prev_btn = ttk.Button(nav_frame, text="\u25C0 Prev",
                                   command=self._prev)
        self.prev_btn.pack(side=tk.LEFT)

        self.next_btn = ttk.Button(nav_frame, text="Next \u25B6",
                                   command=self._next)
        self.next_btn.pack(side=tk.LEFT, padx=(10, 0))

        # Zone buttons area
        zone_outer = ttk.LabelFrame(self.win, text="Zones", padding=10)
        zone_outer.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

        self.zone_buttons = []
        for i, zone in enumerate(self.zones):
            row = i // 2
            col = i % 2

            color = zone["color"]
            text_color = _text_color_for_bg(color)

            btn_frame = tk.Frame(zone_outer, bg=color, bd=2,
                                 relief=tk.RAISED, cursor="hand2")
            btn_frame.grid(row=row, column=col, padx=5, pady=5,
                           sticky="nsew")

            label = tk.Label(btn_frame, text=zone["name"], bg=color,
                             fg=text_color, font=("Helvetica", 14, "bold"),
                             padx=20, pady=12, cursor="hand2")
            label.pack(fill=tk.BOTH, expand=True)

            btn_frame.bind("<Button-1>",
                           lambda e, z=zone["name"]: self._assign_zone(z))
            label.bind("<Button-1>",
                       lambda e, z=zone["name"]: self._assign_zone(z))

            self.zone_buttons.append((btn_frame, label, zone["name"]))

        zone_outer.columnconfigure(0, weight=1)
        zone_outer.columnconfigure(1, weight=1)
        num_rows = (len(self.zones) + 1) // 2
        for r in range(num_rows):
            zone_outer.rowconfigure(r, weight=1)

        # Bottom bar
        bottom_frame = ttk.Frame(self.win, padding=(15, 5, 15, 10))
        bottom_frame.pack(fill=tk.X)

        ttk.Button(bottom_frame, text="Clear Zone",
                   command=self._clear_zone).pack(side=tk.LEFT)
        ttk.Button(bottom_frame, text="Split Zones",
                   command=self._open_split_zones).pack(side=tk.LEFT, padx=(10, 0))

        self.done_btn = ttk.Button(bottom_frame, text="Done",
                                   command=self._done)
        self.done_btn.pack(side=tk.RIGHT, padx=(5, 0))

        ttk.Button(bottom_frame, text="Cancel",
                   command=self._cancel).pack(side=tk.RIGHT)

    def _show_employee(self):
        if not self.employees:
            return

        emp = self.employees[self.current_index]
        self.name_label.config(text=emp["name"])
        self.info_label.config(
            text=f"{emp.get('shift_text', '')}  |  {emp.get('job', '')}")

        total = len(self.employees)
        zoned = sum(1 for e in self.employees if e.get("zone"))
        self.progress_label.config(
            text=f"Employee {self.current_index + 1} of {total}"
                 f"  ({zoned} zoned)")
        self.progress_bar["maximum"] = total
        self.progress_bar["value"] = zoned

        # Update nav buttons
        self.prev_btn.config(
            state=tk.NORMAL if self.current_index > 0 else tk.DISABLED)
        self.next_btn.config(
            state=tk.NORMAL if self.current_index < total - 1
            else tk.DISABLED)

        # Highlight currently assigned zone (only if single-zone, not split)
        current_zone = emp.get("zone", "")
        has_splits = bool(emp.get("zone_segments"))
        for btn_frame, label, zone_name in self.zone_buttons:
            if zone_name == current_zone and not has_splits:
                btn_frame.config(relief=tk.SUNKEN, bd=3)
            else:
                btn_frame.config(relief=tk.RAISED, bd=2)

        # If all employees are zoned, focus Done button
        if zoned == total:
            self.done_btn.focus_set()

    def _assign_zone(self, zone_name):
        if not self.employees:
            return
        self.employees[self.current_index]["zone"] = zone_name
        self.employees[self.current_index]["zone_segments"] = []

        if self.mode == "single":
            self.win.destroy()
            return

        # Sequential: advance to next employee
        if self.current_index < len(self.employees) - 1:
            self.current_index += 1
        self._show_employee()

    def _prev(self):
        if self.current_index > 0:
            self.current_index -= 1
            self._show_employee()

    def _next(self):
        if self.current_index < len(self.employees) - 1:
            self.current_index += 1
            self._show_employee()

    def _clear_zone(self):
        if self.employees:
            self.employees[self.current_index]["zone"] = ""
            self.employees[self.current_index]["zone_segments"] = []
            self._show_employee()

    def _zone_by_number(self, num):
        if num <= len(self.zones):
            self._assign_zone(self.zones[num - 1]["name"])

    def _done(self):
        self.win.destroy()

    def _open_split_zones(self):
        if not self.employees:
            return
        emp = self.employees[self.current_index]
        dialog = SplitZoneDialog(self.win, emp, self.zones)
        if dialog.result is not None:
            emp["zone_segments"] = dialog.result
            # Build summary zone string
            zone_names = []
            for seg in dialog.result:
                if seg["zone"] not in zone_names:
                    zone_names.append(seg["zone"])
            emp["zone"] = " / ".join(zone_names) if zone_names else ""

            if self.mode == "single":
                self.win.destroy()
                return

            # Sequential: advance to next employee
            if self.current_index < len(self.employees) - 1:
                self.current_index += 1
            self._show_employee()

    def _cancel(self):
        self.cancelled = True
        self.win.destroy()


class SplitZoneDialog:
    """Two-phase dialog: Phase 1 adds splits, Phase 2 sequentially assigns zones."""

    def __init__(self, parent, emp, zones):
        self.emp = emp
        self.zones = zones
        self.result = None
        self.current_block_index = 0

        # Build initial blocks from shift segments (respects break gaps)
        self.blocks = []
        for (sh, sm), (eh, em) in emp.get("shift_segments", []):
            self.blocks.append({
                "start": (sh, sm), "end": (eh, em), "zone": ""
            })

        # Pre-populate from existing zone_segments if any
        existing = emp.get("zone_segments", [])
        if existing:
            self.blocks = []
            for seg in existing:
                self.blocks.append({
                    "start": tuple(seg["start"]),
                    "end": tuple(seg["end"]),
                    "zone": seg["zone"],
                })

        # Store original blocks for "Remove All Splits"
        self.original_blocks = []
        for (sh, sm), (eh, em) in emp.get("shift_segments", []):
            self.original_blocks.append({
                "start": (sh, sm), "end": (eh, em), "zone": ""
            })

        self.win = tk.Toplevel(parent)
        self.win.title("Split Zones")
        self.win.geometry("650x580")
        self.win.transient(parent)
        self.win.grab_set()

        self._build_ui()
        self._show_phase1()

        self.win.protocol("WM_DELETE_WINDOW", self._cancel)
        self.win.bind("<Escape>", lambda e: self._cancel())
        self.win.wait_window()

    def _build_ui(self):
        # Header (shared)
        header = ttk.Frame(self.win, padding=(15, 10, 15, 5))
        header.pack(fill=tk.X)

        ttk.Label(header, text=self.emp["name"],
                  font=("Helvetica", 20, "bold")).pack(anchor=tk.W)
        ttk.Label(header, text=self.emp.get("shift_text", ""),
                  font=("Helvetica", 14), foreground="gray").pack(anchor=tk.W)

        # Blocks treeview (shared)
        tree_frame = ttk.LabelFrame(self.win, text="Time Blocks", padding=8)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=(5, 5))

        cols = ("num", "time", "zone")
        self.block_tree = ttk.Treeview(tree_frame, columns=cols,
                                       show="headings", height=6,
                                       selectmode="browse")
        self.block_tree.heading("num", text="#")
        self.block_tree.heading("time", text="Time Range")
        self.block_tree.heading("zone", text="Zone")
        self.block_tree.column("num", width=40, anchor=tk.CENTER)
        self.block_tree.column("time", width=200, anchor=tk.CENTER)
        self.block_tree.column("zone", width=200, anchor=tk.CENTER)
        self.block_tree.pack(fill=tk.BOTH, expand=True)

        # ---- Phase 1: Split Controls ----
        self.phase1_frame = ttk.Frame(self.win)

        split_frame = ttk.LabelFrame(self.phase1_frame,
                                     text="Split Controls", padding=8)
        split_frame.pack(fill=tk.X, padx=15, pady=(0, 5))

        # Compute overall shift range in minutes for the slider
        segments = self.emp.get("shift_segments", [])
        if segments:
            self._slider_start = segments[0][0][0] * 60 + segments[0][0][1]
            self._slider_end = segments[-1][1][0] * 60 + segments[-1][1][1]
        else:
            self._slider_start = 0
            self._slider_end = 1440

        # Timeline canvas with tick marks
        self._track_pad_left = 18
        self._track_pad_right = 18
        self.timeline_canvas = tk.Canvas(split_frame, height=32,
                                         highlightthickness=0)
        self.timeline_canvas.pack(fill=tk.X, pady=(0, 0))
        self.timeline_canvas.bind("<Configure>", self._draw_timeline)

        # Slider
        self.slider_var = tk.IntVar(value=self._slider_start + 30)
        self.split_slider = ttk.Scale(
            split_frame, from_=self._slider_start,
            to=self._slider_end,
            orient=tk.HORIZONTAL,
            variable=self.slider_var,
            command=self._on_slider_change,
        )
        self.split_slider.pack(fill=tk.X, pady=(0, 8))
        self._on_slider_change(None)

        btn_row = ttk.Frame(split_frame)
        btn_row.pack(fill=tk.X)

        tk.Button(btn_row, text="Add Split",
                  command=self._add_split).pack(side=tk.LEFT, padx=(0, 10))
        tk.Button(btn_row, text="Remove Last Split",
                  command=self._remove_last_split).pack(side=tk.LEFT, padx=(0, 10))
        tk.Button(btn_row, text="Remove All Splits",
                  command=self._remove_all_splits).pack(side=tk.LEFT)

        # Phase 1 bottom bar
        p1_bottom = ttk.Frame(self.phase1_frame, padding=(15, 5, 15, 10))
        p1_bottom.pack(fill=tk.X)

        tk.Button(p1_bottom, text="Set Zones \u25B6",
                  command=self._enter_phase2,
                  font=("Helvetica", 12, "bold")).pack(side=tk.RIGHT, padx=(5, 0))
        tk.Button(p1_bottom, text="Cancel",
                  command=self._cancel).pack(side=tk.RIGHT)

        # ---- Phase 2: Zone Assignment ----
        self.phase2_frame = ttk.Frame(self.win)

        # Current block info
        block_info = ttk.Frame(self.phase2_frame, padding=(15, 5))
        block_info.pack(fill=tk.X)

        self.block_label = ttk.Label(block_info, text="",
                                     font=("Helvetica", 18, "bold"))
        self.block_label.pack(anchor=tk.W)

        self.block_progress = ttk.Label(block_info, text="",
                                        foreground="gray")
        self.block_progress.pack(anchor=tk.W)

        # Zone buttons
        zone_frame = ttk.LabelFrame(self.phase2_frame, text="Select Zone",
                                    padding=10)
        zone_frame.pack(fill=tk.X, padx=15, pady=(0, 5))

        self.zone_buttons = []
        for i, zone in enumerate(self.zones):
            row = i // 2
            col = i % 2
            color = zone["color"]
            text_color = _text_color_for_bg(color)

            btn_frame = tk.Frame(zone_frame, bg=color, bd=2,
                                 relief=tk.RAISED, cursor="hand2")
            btn_frame.grid(row=row, column=col, padx=4, pady=3, sticky="nsew")

            label = tk.Label(btn_frame, text=zone["name"], bg=color,
                             fg=text_color, font=("Helvetica", 12, "bold"),
                             padx=15, pady=6, cursor="hand2")
            label.pack(fill=tk.BOTH, expand=True)

            btn_frame.bind("<Button-1>",
                           lambda e, z=zone["name"]: self._assign_zone(z))
            label.bind("<Button-1>",
                       lambda e, z=zone["name"]: self._assign_zone(z))

            self.zone_buttons.append((btn_frame, label, zone["name"]))

        zone_frame.columnconfigure(0, weight=1)
        zone_frame.columnconfigure(1, weight=1)

        # Phase 2 bottom bar
        p2_bottom = ttk.Frame(self.phase2_frame, padding=(15, 5, 15, 10))
        p2_bottom.pack(fill=tk.X)

        tk.Button(p2_bottom, text="\u25C0 Back to Splits",
                  command=self._back_to_phase1).pack(side=tk.LEFT)

        self.done_btn = tk.Button(p2_bottom, text="Done",
                                  command=self._done)
        self.done_btn.pack(side=tk.RIGHT, padx=(5, 0))
        tk.Button(p2_bottom, text="Cancel",
                  command=self._cancel).pack(side=tk.RIGHT)

    # ---- Phase switching ----

    def _show_phase1(self):
        self.phase2_frame.pack_forget()
        self.phase1_frame.pack(fill=tk.BOTH, expand=False)
        self._refresh_blocks()

    def _enter_phase2(self):
        if len(self.blocks) < 2:
            messagebox.showinfo(
                "No Splits",
                "Add at least one split before setting zones.",
                parent=self.win)
            return
        # Clear all zones for fresh assignment
        for blk in self.blocks:
            blk["zone"] = ""
        self.current_block_index = 0
        self.phase1_frame.pack_forget()
        self.phase2_frame.pack(fill=tk.BOTH, expand=False)
        self._refresh_blocks()
        self._show_current_block()

    def _back_to_phase1(self):
        self._show_phase1()

    # ---- Phase 1 methods ----

    def _refresh_blocks(self):
        for item in self.block_tree.get_children():
            self.block_tree.delete(item)
        for i, blk in enumerate(self.blocks):
            start_str = format_time_short(*blk["start"])
            end_str = format_time_short(*blk["end"])
            self.block_tree.insert("", tk.END, values=(
                i + 1, f"{start_str} - {end_str}", blk["zone"] or "(none)"
            ))
        # Select current block
        children = self.block_tree.get_children()
        if children:
            idx = min(self.current_block_index, len(children) - 1)
            self.block_tree.selection_set(children[idx])
        self._refresh_slider()

    def _on_slider_change(self, value):
        """Snap slider to nearest 30-min increment and redraw indicator."""
        raw = self.slider_var.get()
        snapped = round(raw / 30) * 30
        snapped = max(self._slider_start + 30,
                      min(snapped, self._slider_end - 30))
        if snapped != raw:
            self.slider_var.set(snapped)
        self._draw_timeline(None)

    def _draw_timeline(self, event):
        """Draw tick marks and time labels on the timeline canvas."""
        c = self.timeline_canvas
        c.delete("all")
        w = c.winfo_width()
        if w < 10:
            return

        pad_l = self._track_pad_left
        pad_r = self._track_pad_right
        track_w = w - pad_l - pad_r
        if track_w <= 0:
            return

        total_range = self._slider_end - self._slider_start
        if total_range <= 0:
            return

        current_val = self.slider_var.get()

        # Draw each 30-min tick
        t = self._slider_start
        while t <= self._slider_end:
            frac = (t - self._slider_start) / total_range
            x = pad_l + frac * track_w
            is_hour = (t % 60 == 0)

            tick_h = 12 if is_hour else 7
            c.create_line(x, 32 - tick_h, x, 32, fill="#888888",
                          width=2 if is_hour else 1)

            if is_hour:
                h, m = divmod(t, 60)
                lbl = format_time_short(h, m)
                c.create_text(x, 32 - tick_h - 8, text=lbl,
                              font=("Helvetica", 9), fill="#444444",
                              anchor=tk.S)
            t += 30

        # Current position indicator
        cur_frac = (current_val - self._slider_start) / total_range
        cx = pad_l + cur_frac * track_w
        h, m = divmod(current_val, 60)
        c.create_text(cx, 2, text=format_time_short(h, m),
                      font=("Helvetica", 11, "bold"), fill="#0066CC",
                      anchor=tk.N)

    def _refresh_slider(self):
        self._draw_timeline(None)

    def _add_split(self):
        split_min = self.slider_var.get()
        split_time = divmod(split_min, 60)

        for blk in self.blocks:
            if blk["start"] == split_time or blk["end"] == split_time:
                messagebox.showinfo(
                    "Split", f"{format_time_short(*split_time)} is already "
                    "a block boundary.", parent=self.win)
                return

        for i, blk in enumerate(self.blocks):
            blk_start = blk["start"][0] * 60 + blk["start"][1]
            blk_end = blk["end"][0] * 60 + blk["end"][1]
            if blk_start < split_min < blk_end:
                new_block1 = {
                    "start": blk["start"],
                    "end": split_time,
                    "zone": blk["zone"],
                }
                new_block2 = {
                    "start": split_time,
                    "end": blk["end"],
                    "zone": "",
                }
                self.blocks[i:i+1] = [new_block1, new_block2]
                self.current_block_index = i + 1
                self._refresh_blocks()
                return

        messagebox.showinfo("Split",
                            "That time falls in a break gap, not a shift block.",
                            parent=self.win)

    def _remove_last_split(self):
        if len(self.blocks) <= len(self.original_blocks):
            messagebox.showinfo("Remove Split",
                                "No splits to remove.", parent=self.win)
            return
        last = self.blocks.pop()
        self.blocks[-1]["end"] = last["end"]
        self.current_block_index = max(0, len(self.blocks) - 1)
        self._refresh_blocks()

    def _remove_all_splits(self):
        self.blocks = []
        for blk in self.original_blocks:
            self.blocks.append({
                "start": blk["start"],
                "end": blk["end"],
                "zone": "",
            })
        self.current_block_index = 0
        self._refresh_blocks()

    # ---- Phase 2 methods ----

    def _show_current_block(self):
        blk = self.blocks[self.current_block_index]
        start_str = format_time_short(*blk["start"])
        end_str = format_time_short(*blk["end"])
        self.block_label.config(
            text=f"Block {self.current_block_index + 1}: "
                 f"{start_str} - {end_str}")

        zoned = sum(1 for b in self.blocks if b["zone"])
        self.block_progress.config(
            text=f"Block {self.current_block_index + 1} of {len(self.blocks)}"
                 f"  ({zoned}/{len(self.blocks)} zoned)")

        # Highlight the current block in treeview
        children = self.block_tree.get_children()
        if children:
            self.block_tree.selection_set(
                children[self.current_block_index])

        # Highlight assigned zone button
        current_zone = blk.get("zone", "")
        for btn_frame, label, zone_name in self.zone_buttons:
            if zone_name == current_zone and current_zone:
                btn_frame.config(relief=tk.SUNKEN, bd=3)
            else:
                btn_frame.config(relief=tk.RAISED, bd=2)

    def _assign_zone(self, zone_name):
        self.blocks[self.current_block_index]["zone"] = zone_name
        self._refresh_blocks()

        # Auto-advance to next unzoned block
        if self.current_block_index < len(self.blocks) - 1:
            self.current_block_index += 1
        self._show_current_block()

        # If all blocks zoned, focus Done
        if all(b["zone"] for b in self.blocks):
            self.done_btn.focus_set()

    def _done(self):
        for i, blk in enumerate(self.blocks):
            if not blk["zone"]:
                messagebox.showwarning(
                    "Missing Zone",
                    f"Block {i + 1} has no zone assigned.",
                    parent=self.win)
                return
        self.result = [
            {"zone": blk["zone"],
             "start": blk["start"],
             "end": blk["end"]}
            for blk in self.blocks
        ]
        self.win.destroy()

    def _cancel(self):
        self.result = None
        self.win.destroy()


class EmployeeEditDialog:
    def __init__(self, parent, emp):
        self.result = None
        self.win = tk.Toplevel(parent)
        self.win.title("Edit Employee" if emp else "Add Employee")
        self.win.geometry("400x230")
        self.win.transient(parent)
        self.win.grab_set()

        fields = ttk.Frame(self.win)
        fields.pack(padx=10, pady=10, fill=tk.X)

        ttk.Label(fields, text="Name:").grid(row=0, column=0,
                                              sticky=tk.W, pady=2)
        self.name_var = tk.StringVar(value=emp["name"] if emp else "")
        ttk.Entry(fields, textvariable=self.name_var, width=35).grid(
            row=0, column=1, pady=2)

        ttk.Label(fields, text="Shift (e.g. 9A-5P):").grid(
            row=1, column=0, sticky=tk.W, pady=2)
        self.shift_var = tk.StringVar(
            value=emp.get("shift_text", "") if emp else "")
        ttk.Entry(fields, textvariable=self.shift_var, width=35).grid(
            row=1, column=1, pady=2)

        ttk.Label(fields, text="Break (e.g. 12P):").grid(
            row=2, column=0, sticky=tk.W, pady=2)
        self.break_var = tk.StringVar(
            value=emp.get("break_text", "") if emp else "")
        ttk.Entry(fields, textvariable=self.break_var, width=35).grid(
            row=2, column=1, pady=2)

        ttk.Label(fields, text="Job:").grid(
            row=3, column=0, sticky=tk.W, pady=2)
        self.job_var = tk.StringVar(
            value=emp.get("job", "") if emp else "")
        ttk.Entry(fields, textvariable=self.job_var, width=35).grid(
            row=3, column=1, pady=2)

        ttk.Label(fields, text="Zone:").grid(
            row=4, column=0, sticky=tk.W, pady=2)
        self.zone_var = tk.StringVar(
            value=emp.get("zone", "") if emp else "")
        ttk.Entry(fields, textvariable=self.zone_var, width=35).grid(
            row=4, column=1, pady=2)

        btn_frame = ttk.Frame(self.win)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="OK", command=self._ok).pack(
            side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel",
                   command=self.win.destroy).pack(side=tk.LEFT, padx=5)

        self.win.wait_window()

    def _ok(self):
        name = self.name_var.get().strip()
        if not name:
            messagebox.showwarning("Warning", "Name cannot be empty.",
                                   parent=self.win)
            return

        shift_text = self.shift_var.get().strip()

        # Parse shift segments from text like "9A-5P" or "10A-2P / 3P-7P"
        segments = []
        if shift_text:
            for part in re.split(r'\s*/\s*', shift_text):
                m = re.match(
                    r'(\d{1,4}(?::\d{2})?\s*(?:AM|PM|A|P))\s*[-–]\s*'
                    r'(\d{1,4}(?::\d{2})?\s*(?:AM|PM|A|P))',
                    part.strip(), re.IGNORECASE
                )
                if m:
                    start = _parse_compact_time(m.group(1))
                    end = _parse_compact_time(m.group(2))
                    if start and end:
                        segments.append((start, end))

        shift_start = segments[0][0] if segments else None
        shift_end = segments[-1][1] if segments else None

        self.result = {
            "name": name,
            "job": self.job_var.get().strip(),
            "shift_text": shift_text,
            "shift_segments": segments,
            "shift_start": shift_start,
            "shift_end": shift_end,
            "break_text": self.break_var.get().strip(),
            "zone": self.zone_var.get().strip(),
            "zone_segments": [],
        }
        self.win.destroy()


def _parse_compact_time(text):
    """Parse compact time like '9A', '930A', '12:30 PM' into (hour, minute)."""
    if not text:
        return None
    text = text.strip().upper().replace(".", "")

    # Try HH:MM AM/PM
    m = re.match(r'(\d{1,2}):(\d{2})\s*(AM|PM|A|P)', text)
    if m:
        hour, minute = int(m.group(1)), int(m.group(2))
        ap = m.group(3)[0]
        if ap == "P" and hour != 12:
            hour += 12
        if ap == "A" and hour == 12:
            hour = 0
        return (hour, minute)

    # Compact: 930A, 12P, etc.
    m = re.match(r'(\d{1,4})\s*(A|P|AM|PM)', text)
    if m:
        num_str = m.group(1)
        ap = m.group(2)[0]
        if len(num_str) <= 2:
            hour, minute = int(num_str), 0
        elif len(num_str) == 3:
            hour, minute = int(num_str[0]), int(num_str[1:])
        else:
            hour, minute = int(num_str[:2]), int(num_str[2:])
        if ap == "P" and hour != 12:
            hour += 12
        if ap == "A" and hour == 12:
            hour = 0
        return (hour, minute)

    return None


class PreviewDialog:
    def __init__(self, parent, text):
        self.win = tk.Toplevel(parent)
        self.win.title("Playbook Preview")
        self.win.geometry("650x400")
        self.win.transient(parent)

        txt = tk.Text(self.win, font=("Courier", 11), wrap=tk.NONE)
        txt.insert("1.0", text)
        txt.configure(state=tk.DISABLED)

        y_scroll = ttk.Scrollbar(self.win, orient=tk.VERTICAL,
                                  command=txt.yview)
        txt.configure(yscrollcommand=y_scroll.set)
        txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        y_scroll.pack(side=tk.LEFT, fill=tk.Y)

        ttk.Button(self.win, text="Close",
                   command=self.win.destroy).pack(pady=5)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if Workbook is None:
        print("Missing dependency: openpyxl")
        print("Install with: pip install openpyxl")

    root = tk.Tk()
    DCOCreatorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
